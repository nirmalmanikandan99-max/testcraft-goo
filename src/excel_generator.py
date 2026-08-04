from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)


def _to_cell_text(value):
    """
    Coerce a field value into something Excel can store.

    Models sometimes return a dict or list for a field (e.g. Test Data
    as {"Username": "x", "Password": "y"}). openpyxl can only store
    scalars, so flatten structures into readable text.
    """

    if value is None:
        return ""

    if isinstance(value, dict):
        return "\n".join(f"{k}: {_to_cell_text(v)}" for k, v in value.items())

    if isinstance(value, (list, tuple)):
        return "\n".join(_to_cell_text(item) for item in value)

    if isinstance(value, bool):
        return "Yes" if value else "No"

    return str(value)

# Column layouts per format. Each entry is (Header shown in Excel,
# key expected in the test-case dict from the LLM).
CONVENTIONAL_COLUMNS = [
    ("S.No", "S.No"),
    ("Title of Test Case", "Title of Test Case"),
    ("Pre Requisites", "Pre Requisites"),
    ("Actions to be done", "Actions to be done"),
    ("Expected Results", "Expected Results"),
    ("Test Data", "Test Data"),
    ("Testing Technique", "Testing Technique"),
]

GWT_COLUMNS = [
    ("S.No", "S.No"),
    ("Title", "Title"),
    ("Given", "Given"),
    ("When", "When"),
    ("Then", "Then"),
    ("Test Data", "Test Data"),
    ("Testing Technique", "Testing Technique"),
]


def _columns_for(test_case_format):
    if test_case_format == "GWT":
        return GWT_COLUMNS
    return CONVENTIONAL_COLUMNS


def generate_excel(test_cases, output_file, test_case_format="Conventional Test Case"):

    columns = _columns_for(test_case_format)

    wb = Workbook()

    ws = wb.active
    ws.title = "Test Cases"

    # =====================================================
    # Styles
    # =====================================================

    header_fill = PatternFill(
        start_color="6366F1",
        end_color="6366F1",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
        size=11
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    data_alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # =====================================================
    # Headers
    # =====================================================

    for col, (header, _key) in enumerate(columns, start=1):

        cell = ws.cell(row=1, column=col)

        cell.value = header

        cell.font = header_font

        cell.fill = header_fill

        cell.border = thin_border

        cell.alignment = header_alignment

    # =====================================================
    # Write Test Cases
    # =====================================================

    row = 2

    for index, tc in enumerate(test_cases, start=1):

        # Some models omit S.No; fall back to the row index.
        for col, (_header, key) in enumerate(columns, start=1):

            if key == "S.No":
                value = tc.get(key, index)
            else:
                value = tc.get(key, "")

            cell = ws.cell(row=row, column=col)
            cell.value = _to_cell_text(value)
            cell.border = thin_border
            cell.alignment = data_alignment

        row += 1

    # =====================================================
    # Freeze Header
    # =====================================================

    ws.freeze_panes = "A2"

    # =====================================================
    # Auto Filter
    # =====================================================

    ws.auto_filter.ref = ws.dimensions

    # =====================================================
    # Auto Column Width
    # =====================================================

    for column_cells in ws.columns:

        length = 0

        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            try:

                if cell.value:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

            except Exception:

                pass

        adjusted_width = min(length + 5, 60)

        ws.column_dimensions[column_letter].width = adjusted_width

    # =====================================================
    # Row Height
    # =====================================================

    ws.row_dimensions[1].height = 30

    # =====================================================
    # Save Workbook
    # =====================================================

    wb.save(output_file)

    print("=" * 60)
    print("Excel Generated Successfully")
    print(output_file)
    print("=" * 60)
